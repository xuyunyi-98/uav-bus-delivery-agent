"""
生成50个真实感包裹数据 Excel 文件
运行方式：python generate_orders_excel.py
输出文件：UAV_Bus_Agent/配送包裹数据_50条.xlsx
"""
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

random.seed(42)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "配送订单"

# 样式
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
normal_font = Font(name="微软雅黑", size=10)
thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
center_align = Alignment(horizontal="center", vertical="center")

# 列宽
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 8
ws.column_dimensions["D"].width = 8
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 10

# 标题
ws.merge_cells("A1:G1")
ws["A1"].value = "🚁 农村公交-无人机协同配送 — 包裹订单数据"
ws["A1"].font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

ws.merge_cells("A2:G2")
ws["A2"].value = f"生成日期：{datetime.now().strftime('%Y-%m-%d %H:%M')}  |  共 50 条订单 | 可直接导入 Streamlit 智能体"
ws["A2"].font = Font(name="微软雅黑", size=9, italic=True, color="888888")
ws["A2"].alignment = Alignment(horizontal="center")
ws.row_dimensions[2].height = 20

# 表头
headers = ["id", "x", "y", "weight", "urgency", "item_type", "deadline"]
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[4].height = 25

# 物品类型池（带权重，更真实）
item_types = [
    ("农产品", 0.25), ("日用品", 0.20), ("医疗物资", 0.10),
    ("生鲜食品", 0.12), ("化肥", 0.08), ("快递包裹", 0.10),
    ("饲料", 0.05), ("建材", 0.05), ("农机零件", 0.03), ("药品", 0.02)
]

# 生成50条数据
orders = []
for i in range(1, 51):
    # 坐标：模拟农村分布——近公交线(Y≈50)的包裹较多、偏远也有
    # 80%概率分布在公交服务站30km半径内(Y在20-80之间)
    if random.random() < 0.8:
        x = random.randint(5, 95)
        y = random.randint(20, 80)
    else:
        x = random.randint(0, 100)
        y = random.randint(0, 100)

    # 重量分布：轻量包裹(<5kg)占比60%，中等(5-10kg)30%，超重(>10kg)10%
    r = random.random()
    if r < 0.60:
        weight = round(random.uniform(0.3, 4.9), 1)
    elif r < 0.90:
        weight = round(random.uniform(5.0, 9.9), 1)
    else:
        weight = round(random.uniform(10.1, 20.0), 1)

    # 紧急程度：15%加急，85%普通
    urgency = "加急" if random.random() < 0.15 else "普通"

    # 时间窗：30% 的包裹有最晚送达时间（30~120分钟）
    deadline = None
    if random.random() < 0.30:
        deadline = random.randint(30, 120)

    item_type = random.choices(
        [t[0] for t in item_types],
        weights=[t[1] for t in item_types]
    )[0]

    orders.append({
        "id": i, "x": x, "y": y, "weight": weight,
        "urgency": urgency, "item_type": item_type
    })
    if deadline is not None:
        orders[-1]["deadline"] = deadline

# 写入数据
for row_idx, order in enumerate(orders, 5):
    vals = [order["id"], order["x"], order["y"], order["weight"],
            order["urgency"], order["item_type"], order.get("deadline", "")]
    for col_idx, val in enumerate(vals, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
        # 加急件标红背景
        if order["urgency"] == "加急":
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        # 超重标黄
        if order["weight"] > 10:
            cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    ws.row_dimensions[row_idx].height = 22

# 冻结表头
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:G{4+50}"

# 统计
row = 56
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
ws.cell(row=row, column=1).value = (
    f"📊 共50条 | 加急{sum(1 for o in orders if o['urgency']=='加急')}条 | "
    f"超重(>10kg){sum(1 for o in orders if o['weight']>10)}条 | "
    f"可无人机{sum(1 for o in orders if o['weight']<=10)}条 | "
    f"有时限{sum(1 for o in orders if o.get('deadline'))}条"
)
ws.cell(row=row, column=1).font = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
ws.cell(row=row, column=1).fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

output_path = "UAV_Bus_Agent/配送包裹数据_50条.xlsx"
wb.save(output_path)
print(f"✅ 已生成：{output_path}")
print(f"   - 加急件: {sum(1 for o in orders if o['urgency']=='加急')} 条")
print(f"   - 超重件(>10kg): {sum(1 for o in orders if o['weight']>10)} 条")
print(f"   - 无人机可配送: {sum(1 for o in orders if o['weight']<=10)} 条")