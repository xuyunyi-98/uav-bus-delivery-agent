"""导出模块：一键生成 Excel 派单结果 + PDF 报告"""
import io, os, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_excel(command, weather, objective, metrics, score,
                 merged_df, weights, plan, tasks):
    """生成派单结果 Excel，返回 bytes"""
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "派单结果"

    hdr_font = Font("Microsoft YaHei", 11, bold=True, color="FFFFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1a56db")
    card_fill = PatternFill("solid", fgColor="f0f5ff")
    green_fill = PatternFill("solid", fgColor="e6ffe6")
    red_fill = PatternFill("solid", fgColor="ffe6e6")
    border = Border(left=Side("thin","cccccc"),right=Side("thin","cccccc"),
                    top=Side("thin","cccccc"),bottom=Side("thin","cccccc"))
    center = Alignment(horizontal="center", vertical="center")

    for col, w in {"A":6,"B":10,"C":10,"D":10,"E":10,"F":14,"G":14,"H":14,"I":10,"J":10}.items():
        ws.column_dimensions[col].width = w

    # 标题
    ws.merge_cells("A1:J1")
    c = ws["A1"]; c.value = "🚁 公交-无人机协同配送 · 派单结果报告"
    c.font = Font("Microsoft YaHei", 16, bold=True, color="1a56db")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"{datetime.datetime.now():%Y-%m-%d %H:%M}  |  指令: {command}  |  天气: {weather}  |  目标: {objective}"
    ws["A2"].font = Font("Microsoft YaHei", 9, color="FF888888")
    ws["A2"].alignment = center

    # 指标卡片
    labels = ["包裹数","Fitness","耗时(min)","成本(¥)","碳排(kg)","仓库","无人机","公交","权重(cost)","权重(time)"]
    values = [len(plan), score, metrics["time"], metrics["cost"], metrics["carbon"],
              2, 3, 2, weights.get("cost_weight",0.33), weights.get("time_weight",0.33)]
    for i,(lbl,val) in enumerate(zip(labels,values)):
        c=ws.cell(row=4,column=i+1,value=lbl)
        c.font=Font("Microsoft YaHei",9,bold=True,color="1a56db");c.fill=card_fill;c.alignment=center;c.border=border
        c2=ws.cell(row=5,column=i+1,value=val)
        c2.font=Font("Microsoft YaHei",13,bold=True);c2.fill=card_fill;c2.alignment=center;c2.border=border
    ws.row_dimensions[4].height=22; ws.row_dimensions[5].height=30

    # 明细表
    for i,h in enumerate(["包裹ID","配送方式","车辆编号","仓库","重量(kg)","紧急程度","物品类型"]):
        c=ws.cell(row=7,column=i+1,value=h);c.font=hdr_font;c.fill=hdr_fill;c.alignment=center;c.border=border
    ws.row_dimensions[7].height=24
    for ri,item in enumerate(plan):
        r=8+ri; t=next((x for x in tasks if x["id"]==item["task_id"]),{})
        vals=[item["task_id"],item["method"],item.get("vehicle_id",""),item.get("depot_id",""),
              t.get("weight",""),t.get("urgency",""),t.get("item_type","")]
        fill=red_fill if item["method"]=="无人机" else green_fill
        for ci,v in enumerate(vals):
            c=ws.cell(row=r,column=ci+1,value=v);c.font=Font("Microsoft YaHei",10)
            c.alignment=center;c.border=border;c.fill=fill
        ws.row_dimensions[r].height=22

    sr=8+len(plan)+1
    ws.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=7)
    un=sum(1 for p in plan if p["method"]=="无人机")
    bn=sum(1 for p in plan if p["method"]=="公交")
    ws.cell(row=sr,column=1).value=f"📊 无人机配送 {un} 件 | 公交配送 {bn} 件 | 总计 {len(plan)} 件"
    ws.cell(row=sr,column=1).font=Font("Microsoft YaHei",10,bold=True,color="FF333333")
    ws.cell(row=sr,column=1).fill=PatternFill("solid",fgColor="e8f0fe")

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def export_pdf(command, weather, objective, metrics, score,
               merged_df, weights, plan, tasks):
    """生成 PDF 报告，返回 bytes (需要 reportlab)"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.colors import HexColor
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle, HRFlowable)
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=20*mm, bottomMargin=15*mm,
                            leftMargin=15*mm, rightMargin=15*mm)

    # 中文字体
    font_name = "Helvetica"
    for fp in ["C:/Windows/Fonts/msyh.ttc","C:/Windows/Fonts/simhei.ttf"]:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("Chinese", fp))
                font_name = "Chinese"; break
            except: pass

    styles = getSampleStyleSheet()
    ts = ParagraphStyle("CTitle",parent=styles["Title"],fontName=font_name,
                        fontSize=20,textColor=HexColor("#1a56db"),spaceAfter=6)
    h2s = ParagraphStyle("CH2",parent=styles["Heading2"],fontName=font_name,
                         fontSize=14,textColor=HexColor("#1a56db"),spaceBefore=12)
    ns = ParagraphStyle("CN",parent=styles["Normal"],fontName=font_name,
                        fontSize=10,leading=16)
    ss = ParagraphStyle("CS",parent=styles["Normal"],fontName=font_name,
                        fontSize=9,textColor=HexColor("#666"))

    elems = []
    elems.append(Paragraph("公交-无人机协同配送报告", ts))
    elems.append(Paragraph(f"生成: {datetime.datetime.now():%Y-%m-%d %H:%M}  |  "
                           f"指令: {command}  |  天气: {weather}  |  目标: {objective}", ss))
    elems.append(Spacer(1, 10))

    # 指标表
    cd = [["包裹","Fitness","耗时(min)","成本","碳排","天气","目标"],
          [str(len(plan)),str(score),str(metrics['time']),
           f"¥{metrics['cost']}",f"{metrics['carbon']}kg",weather,objective]]
    ct = Table(cd, colWidths=[45,55,55,55,55,65,80])
    ct.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),HexColor("#1a56db")),
        ("TEXTCOLOR",(0,0),(-1,0),HexColor("#fff")),
        ("FONTNAME",(0,0),(-1,-1),font_name),
        ("FONTSIZE",(0,0),(-1,0),9),("FONTSIZE",(0,1),(-1,-1),12),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("GRID",(0,0),(-1,-1),0.5,HexColor("#ccc")),
        ("BACKGROUND",(0,1),(-1,1),HexColor("#f0f5ff")),
    ]))
    elems.append(ct); elems.append(Spacer(1, 14))

    # 明细表
    elems.append(Paragraph("派单明细", h2s)); elems.append(Spacer(1, 6))
    dd = [["ID","方式","车辆编号","重量","紧急","物品"]]
    for item in plan:
        t = next((x for x in tasks if x["id"]==item["task_id"]),{})
        dd.append([str(item["task_id"]),item["method"],
                   item.get("vehicle_id",""),str(t.get("weight","")),
                   t.get("urgency",""),t.get("item_type","")])

    dt = Table(dd, colWidths=[30,45,110,50,45,75], repeatRows=1)
    cmds = [("BACKGROUND",(0,0),(-1,0),HexColor("#1a56db")),
            ("TEXTCOLOR",(0,0),(-1,0),HexColor("#fff")),
            ("FONTNAME",(0,0),(-1,-1),font_name),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("GRID",(0,0),(-1,-1),0.3,HexColor("#ddd"))]
    for r in range(1, len(dd)):
        bg = HexColor("#ffe6e6") if dd[r][1] == "无人机" else HexColor("#e6ffe6")
        cmds.append(("BACKGROUND",(0,r),(-1,r),bg))
    dt.setStyle(TableStyle(cmds))
    elems.append(dt); elems.append(Spacer(1, 10))

    un = sum(1 for p in plan if p["method"]=="无人机")
    bn = len(plan)-un
    elems.append(Paragraph(f"无人机配送 {un} 件 | 公交配送 {bn} 件 | 总计 {len(plan)} 件", ns))
    elems.append(Spacer(1, 8))
    elems.append(HRFlowable(width="100%", thickness=1, color=HexColor("#ccc")))
    elems.append(Paragraph(
        f"权重: time={weights.get('time_weight',0.33)} "
        f"cost={weights.get('cost_weight',0.33)} "
        f"carbon={weights.get('carbon_weight',0.34)}", ss))

    doc.build(elems); buf.seek(0)
    return buf.getvalue()