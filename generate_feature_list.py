"""
生成智能体功能清单 Excel 表格
运行方式：python generate_feature_list.py
输出文件：UAV_Bus_Agent/智能体功能清单.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ==========================================
# 样式定义
# ==========================================
title_font = Font(name="微软雅黑", size=16, bold=True, color="1F4E79")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
category_font = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
normal_font = Font(name="微软雅黑", size=10)
done_font = Font(name="微软雅黑", size=10, color="2E7D32")
future_font = Font(name="微软雅黑", size=10, color="E65100")
note_font = Font(name="微软雅黑", size=9, italic=True, color="666666")

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
category_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
done_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
future_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ==========================================
# Sheet 1: 功能清单主表
# ==========================================
ws = wb.active
ws.title = "功能清单"

# 列宽设定
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 42
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 14
ws.column_dimensions["G"].width = 40
ws.column_dimensions["H"].width = 40

# 标题行
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "🚁 农村公交-无人机协同配送智能体 — 功能清单"
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

# 副标题
ws.merge_cells("A2:H2")
ws["A2"].value = f"生成日期：{datetime.now().strftime('%Y-%m-%d %H:%M')}     |     状态：✔ = 已实现     ⏳ = 规划中     ✏️ = 可自行编辑更新"
ws["A2"].font = note_font
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

# 表头
headers = ["序号", "功能模块", "功能描述", "状态", "优先级", "实现难度", "技术细节 / 涉及文件", "备注 / 升级方向"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[4].height = 28

# ==========================================
# 功能数据
# ==========================================
features = [
    # 格式: (模块, 描述, 状态, 优先级, 难度, 技术细节, 备注)

    # ===== 1. 智能意图解析 =====
    ("CATEGORY", "一、LLM 智能意图解析", "", "", "", "", ""),
    (True, "自然语言指令理解", "支持中文自然语言输入，DeepSeek LLM 解析用户配送需求", "✔", "—", "已完成", "llm_agent.py / config.py", "可扩展支持英文、语音输入"),
    (True, "优化目标自动识别", "从指令中自动提取优化偏好：成本最低/时间最短/碳排放最低/多目标平衡", "✔", "—", "已完成", "agent_core.py: 权重自适应映射", ""),
    (True, "配送方式强制约束", "识别'全部用公交''所有用无人机'等强制指令，直接锁定配送方式", "✔", "—", "已完成", "agent_core.py: force_all_bus/force_all_drone", "本次新增功能"),
    (True, "空指令保护", "用户未输入指令时，自动回退到多目标平衡模式", "✔", "—", "已完成", "agent_core.py", ""),

    # ===== 2. 订单管理 =====
    ("CATEGORY", "二、订单数据管理", "", "", "", "", ""),
    (True, "Excel 风格数据编辑", "Streamlit data_editor，支持双击修改、增删订单行", "✔", "—", "已完成", "app.py: st.data_editor", ""),
    (True, "订单属性支持", "订单包含：坐标(x,y)、重量(kg)、紧急程度、物品类型", "✔", "—", "已完成", "app.py 默认订单 / task_generator.py", ""),
    (True, "超重自动拦截", "重量 >10kg 的包裹自动判定无人机不可用，分配公交", "✔", "—", "已完成", "environment.py: max_drone_load / fitness.py", ""),
    (False, "批量 Excel 导入", "支持从 .xlsx/.csv 文件直接导入订单数据", "⏳", "高", "低", "app.py 扩展: st.file_uploader + pandas", "提升业务实操效率"),
    (False, "数据库持久化", "订单数据存入 MySQL/PostgreSQL，支持历史查询", "⏳", "高", "中", "新增 db_manager.py / SQLAlchemy", "替代当前纯内存模式"),

    # ===== 3. 核心算法 =====
    ("CATEGORY", "三、EAPSO-HDTA 核心算法", "", "", "", "", ""),
    (True, "改进粒子群优化", "20粒子×20迭代的 EAPSO 算法，含自适应惯性权重", "✔", "—", "已完成", "eapso_agent.py", "专利核心"),
    (True, "多维风险势场 U", "构建无人机偏离度势场，自适应调控搜索方向", "✔", "—", "已完成", "eapso_agent.py: calculate_risk_U()", "专利步骤 S2"),
    (True, "动态时空引力场", "Fatt 引力场引导粒子向公交/无人机合理分配", "✔", "—", "已完成", "eapso_agent.py: get_Fatt()", "专利权利要求 2"),
    (True, "Sigmoid 非线性映射", "k(U) = kmin + (kmax-kmin)/(1+e^(-α(U-Uth)))", "✔", "—", "已完成", "eapso_agent.py: 自适应权重计算", "专利步骤 S3"),
    (True, "极值逃逸变异", "连续2代停滞时，30%概率触发粒子翻转，跳出局部最优", "✔", "—", "已完成", "eapso_agent.py: 变异扰动机制", ""),
    (True, "精英保留策略", "最优粒子不参与变异，保障收敛稳定性", "✔", "—", "已完成", "eapso_agent.py: range(1, particles)", ""),
    (False, "参数自适应调优", "根据问题规模自动调整粒子数/迭代数", "⏳", "中", "中", "eapso_agent.py 动态参数", ""),
    (False, "并行化加速", "多进程/GPU 加速粒子群评估", "⏳", "低", "高", "multiprocessing / CuPy", "大规模问题需要"),

    # ===== 4. 适应度评价 =====
    ("CATEGORY", "四、多目标适应度评价", "", "", "", "", ""),
    (True, "三维度综合评价", "时间(min) + 成本(¥) + 碳排(kg) 加权求和", "✔", "—", "已完成", "fitness.py: calculate()", ""),
    (True, "动态权重映射", "LLM 意图 → 权重自动分配 (如'成本最低'→cost=0.6)", "✔", "—", "已完成", "agent_core.py 权重映射", ""),
    (True, "加急件惩罚机制", "加急件分配给公交 → +1000 惩罚，强制使用无人机", "✔", "—", "已完成", "fitness.py: urgency==加急 → cost+1000", ""),
    (True, "超重物理约束", ">10kg 包裹分给无人机 → +1000 惩罚", "✔", "—", "已完成", "fitness.py: weight>10 → cost+1000", ""),
    (True, "指标明细分解", "返回 time/cost/carbon 三项独立数值供前端展示", "✔", "—", "已完成", "fitness.py: get_details()", ""),
    (False, "更多评价维度", "客户满意度、道路拥堵度、能耗精确模型", "⏳", "中", "中", "fitness.py 扩展", ""),

    # ===== 5. 环境建模 =====
    ("CATEGORY", "五、配送环境建模", "", "", "", "", ""),
    (True, "公交线路模拟", "水平公交线路：3个站点 (0,50)-(50,50)-(100,50)", "✔", "—", "已完成", "environment.py", ""),
    (True, "任务-公交距离计算", "欧式距离计算每个包裹到最近公交站的距离", "✔", "—", "已完成", "environment.py: nearest_bus_distance()", ""),
    (True, "无人机电量模拟", "随机初始化电量(15-100%)，低电量(<20%)自动不可用", "✔", "—", "已完成", "environment.py: drone_batteries", ""),
    (True, "电量实时展示", "侧边栏可视化展示无人机电量状态（红/黄/绿）", "✔", "—", "已完成", "app.py: 电量状态组件", ""),
    (False, "真实地图集成", "接入高德/百度地图 API，使用真实路网和 GPS 坐标", "⏳", "高", "高", "新增 map_integration.py", "从仿真走向真实落地的关键"),
    (False, "多公交线路", "支持多条公交线路、多换乘站", "⏳", "中", "中", "environment.py: bus_routes 列表", ""),
    (False, "动态天气影响", "实时天气 API → 影响无人机能耗/速度", "⏳", "中", "中", "新增 weather_service.py", ""),

    # ===== 6. 车辆调度 =====
    ("CATEGORY", "六、车辆调度与编号", "", "", "", "", ""),
    (True, "运力池设定", "前端侧边栏手动设定可用无人机/公交数量", "✔", "—", "已完成", "app.py: st.number_input", ""),
    (True, "车辆编号分配", "每个包裹分配具体车辆：UAV-1, UAV-2, Bus-1, Bus-2...", "✔", "—", "已完成", "eapso_agent.py: _assign_vehicle_ids()", "本次新增功能"),
    (True, "轮询负载均衡", "包裹按轮询方式均匀分配给同类型车辆", "✔", "—", "已完成", "eapso_agent.py: 轮询策略", ""),
    (False, "车辆路径规划", "单车辆多包裹的最优配送顺序 (TSP)", "⏳", "高", "高", "新增 routing.py / OR-Tools", "提升实际配送效率"),
    (False, "时间窗约束", "包裹有最晚送达时间限制", "⏳", "中", "中", "fitness.py 时间窗惩罚", ""),
    (False, "动态车辆增减", "运行中动态上线/下线车辆", "⏳", "低", "中", "agent_core.py 动态运力池", ""),

    # ===== 7. 可视化 =====
    ("CATEGORY", "七、可视化分析", "", "", "", "", ""),
    (True, "配送规划图", "红色三角(无人机) + 绿色圆(公交)，含车辆编号标注", "✔", "—", "已完成", "visualization.py", "自动保存至 results/"),
    (True, "算法收敛曲线", "EAPSO 迭代收敛过程实时展示", "✔", "—", "已完成", "app.py: 收敛曲线子图", ""),
    (True, "多算法基准对比", "EAPSO vs Standard PSO vs Traditional GA", "✔", "—", "已完成", "app.py: 基准对比图", ""),
    (True, "结果数据表格", "带颜色标记的派单结果明细表（红=无人机，绿=公交）", "✔", "—", "已完成", "app.py: st.dataframe + 样式", ""),
    (True, "经验记忆库", "历史调度决策记录对比表", "✔", "—", "已完成", "app.py: st.session_state.memory_bank", ""),
    (False, "实时地图热力图", "订单密度热力图 + 车辆实时位置", "⏳", "中", "中", "folium / pydeck", ""),
    (False, "大屏驾驶舱", "全屏实时监控 Dashboard", "⏳", "低", "中", "Streamlit 布局优化", ""),

    # ===== 8. 智能体经验 =====
    ("CATEGORY", "八、智能体持续学习", "", "", "", "", ""),
    (True, "决策记忆存储", "每次调度的指令-结果存入 session 记忆库", "✔", "—", "已完成", "app.py: memory_bank", ""),
    (False, "偏好学习", "从历史决策中学习用户偏好，自动调整默认权重", "⏳", "高", "高", "新增 preference_learner.py", "真正的'智能化'"),
    (False, "异常检测告警", "检测异常调度结果并自动告警", "⏳", "中", "中", "新增 anomaly_detector.py", ""),
    (False, "A/B 策略对比", "同一批订单跑两种策略，对比效果", "⏳", "低", "中", "新增 ab_test.py", ""),

    # ===== 9. 系统集成 =====
    ("CATEGORY", "九、系统集成与部署", "", "", "", "", ""),
    (True, "Streamlit Web 界面", "完整可交互的 Web UI", "✔", "—", "已完成", "app.py (streamlit run)", ""),
    (True, "图片自动保存", "规划图/收敛曲线/对比图自动保存至 results/", "✔", "—", "已完成", "visualization.py / app.py", ""),
    (False, "REST API 接口", "FastAPI/Flask 暴露调度 API", "⏳", "高", "中", "新增 api_server.py", "供外部系统调用"),
    (False, "Docker 容器化", "一键部署的 Docker 镜像", "⏳", "中", "低", "Dockerfile / docker-compose.yml", ""),
    (False, "移动端适配", "响应式布局适配手机/平板", "⏳", "低", "中", "Streamlit 主题配置", ""),
    (False, "PDF 报告导出", "一键导出含图表和数据的 PDF 调度报告", "⏳", "中", "中", "reportlab / weasyprint", ""),

    # ===== 10. 业务扩展 =====
    ("CATEGORY", "十、业务场景扩展", "", "", "", "", ""),
    (False, "多仓库协同", "多个配送中心同时调度", "⏳", "中", "高", "environment.py 多仓库建模", ""),
    (False, "取件+派件", "支持双向物流（取件+派件）", "⏳", "中", "中", "task 增加 type 字段", ""),
    (False, "冷链配送", "生鲜/药品的温度监控与时效保障", "⏳", "低", "中", "fitness.py 冷链惩罚", ""),
    (False, "应急救灾模式", "灾害场景下不考虑成本、只追求时效", "⏳", "中", "中", "新增 emergency_mode", ""),
]

# ==========================================
# 写入数据
# ==========================================
row = 5
seq = 0
category_seq = 0

for item in features:
    if item[0] == "CATEGORY":
        # 分类标题行
        category_seq += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=f"{item[1]}")
        cell.font = category_font
        cell.fill = category_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        for c in range(2, 9):
            ws.cell(row=row, column=c).fill = category_fill
            ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 26
        row += 1
        continue

    is_done = item[0]
    seq += 1

    status = item[2]
    fill = done_fill if status == "✔" else future_fill
    r_font = done_font if status == "✔" else future_font

    data = [seq, item[1], item[2], status, item[4], item[5], item[6], item[7]]
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = r_font if col_idx == 4 else (normal_font if status == "✔" else future_font)
        cell.border = thin_border
        cell.fill = fill
        if col_idx in (1, 4, 5, 6):
            cell.alignment = center_align
        else:
            cell.alignment = left_align

    # 功能描述列加粗
    ws.cell(row=row, column=3).font = Font(name="微软雅黑", size=10, bold=True, color="1F4E79" if status == "✔" else "E65100")

    ws.row_dimensions[row].height = 30 if is_done else 28
    row += 1

# ==========================================
# 统计行
# ==========================================
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
done_count = sum(1 for f in features if f[0] is True)
future_count = sum(1 for f in features if f[0] is False)
cell = ws.cell(row=row, column=1, value=f"📊 统计：已实现 {done_count} 项  |  规划中 {future_count} 项  |  完成率 {done_count}/{done_count+future_count} = {round(done_count/(done_count+future_count)*100)}%")
cell.font = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
cell.border = thin_border
for c in range(2, 9):
    ws.cell(row=row, column=c).fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    ws.cell(row=row, column=c).border = thin_border
ws.row_dimensions[row].height = 26

# ==========================================
# Sheet 2: 使用说明
# ==========================================
ws2 = wb.create_sheet("使用说明")
ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 70

instructions = [
    ("使用说明", True),
    ("", False),
    ("📋 如何更新功能状态：", True),
    ('1. 在「功能清单」表中找到你要更新的功能行', False),
    ('2. 将「状态」列的 ⏳ 改为 ✔（或粘贴 ✔ 符号）', False),
    ("3. 保存 Excel 文件即可", False),
    ("", False),
    ("✏️ 如何添加新功能：", True),
    ("1. 在对应分类下方插入新行", False),
    ("2. 填写功能模块、功能描述、状态(⏳)、优先级、难度、技术细节", False),
    ("3. 状态列推荐使用 ✔（已实现）或 ⏳（规划中）", False),
    ("", False),
    ("🎨 颜色说明：", True),
    ("绿色行 = 已实现的功能", False),
    ("橙色行 = 规划中/待开发的功能", False),
    ("蓝色行 = 功能分类标题", False),
    ("", False),
    ("📊 当前统计（截至代码分析）：", True),
    (f"已实现功能：{done_count} 项", False),
    (f"规划中功能：{future_count} 项", False),
    (f"整体完成率：{round(done_count/(done_count+future_count)*100)}%", False),
    ("", False),
    ("💡 提示：", True),
    ("• 优先级分为：高 / 中 / 低", False),
    ("• 实现难度分为：高 / 中 / 低 / 已完成", False),
    ('• 「备注/升级方向」列可以填写你的个人想法', False),
    ("• 每次迭代后可更新此表，记录版本演进", False),
]

for i, (text, is_title) in enumerate(instructions, 1):
    cell = ws2.cell(row=i, column=2, value=text)
    if is_title:
        cell.font = Font(name="微软雅黑", size=13, bold=True, color="1F4E79")
    else:
        cell.font = Font(name="微软雅黑", size=10, color="333333")
    ws2.row_dimensions[i].height = 22 if not is_title else 30

# 冻结表头
ws.freeze_panes = "A5"

# 添加筛选
ws.auto_filter.ref = f"A4:H{row - 2}"

# 保存
output_path = "UAV_Bus_Agent/智能体功能清单.xlsx"
wb.save(output_path)
print(f"✅ 功能清单已生成：{output_path}")
print(f"   - 已实现功能：{done_count} 项")
print(f"   - 规划中功能：{future_count} 项")
print(f"   - 完成率：{round(done_count/(done_count+future_count)*100)}%")